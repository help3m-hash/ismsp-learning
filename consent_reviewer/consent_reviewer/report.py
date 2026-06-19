"""엑셀 리포트 생성 (openpyxl).

ReviewResult 목록 → '요약' 시트 + 문서별 상세 시트.
상세 시트는 기존 처리방침 점검 도구와 동일한 형식:
  제목/대상파일/점검일자/통계/총평 블록 + 표
  (번호 | 점검 영역 | 세부 점검항목 | 판정 | 근거(인용) | 미흡 사유 | 개선 권고 | 신뢰도 | 근거 법령)
"""

from __future__ import annotations

import os
import re
from collections import Counter
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import config
from .knowledge import get_checklist
from .models import ReviewResult

# 판정별 배경색
_VERDICT_FILL = {
    "적합": PatternFill("solid", fgColor="C6EFCE"),
    "보완필요": PatternFill("solid", fgColor="FFEB9C"),
    "누락": PatternFill("solid", fgColor="FFC7CE"),
    "해당없음": PatternFill("solid", fgColor="E7E6E6"),
}
_CONF_FILL = {
    "높음": PatternFill("solid", fgColor="C6EFCE"),
    "중간": PatternFill("solid", fgColor="FFF2CC"),
    "낮음": PatternFill("solid", fgColor="FCE4D6"),
}
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=14)
_META_FONT = Font(size=10, color="404040")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# (제목, 너비)
_COLUMNS = [
    ("번호", 6),
    ("점검 영역", 18),
    ("세부 점검항목", 26),
    ("판정", 9),
    ("근거(인용)", 34),
    ("미흡 사유", 34),
    ("개선 권고", 34),
    ("신뢰도", 8),
    ("근거 법령", 26),
]
_NCOL = len(_COLUMNS)


def _checklist_index():
    """item_id → (area_no, area, item_name, law_basis, order) 매핑과 영역번호 부여."""
    checklist = get_checklist()
    area_no: dict[str, int] = {}
    idx: dict[str, dict] = {}
    for order, item in enumerate(checklist):
        if item.area not in area_no:
            area_no[item.area] = len(area_no) + 1
        idx[item.item_id] = {
            "area_no": area_no[item.area],
            "area": item.area,
            "item_name": item.item_name,
            "law_basis": item.law_basis,
            "order": order,
        }
    return idx


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """엑셀 시트명 제약(31자, []:*?/\\ 금지, 중복불가)에 맞게 정리."""
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)[:28] or "문서"
    base, n = name, 1
    while name in used:
        suffix = f"_{n}"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def _write_block(ws, result: ReviewResult, counts: Counter, applicable: int) -> int:
    """상세 시트 상단 메타 블록을 쓰고, 표 헤더가 시작될 행 번호를 반환."""
    last_col = get_column_letter(_NCOL)

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = config.REPORT_TITLE
    c.font = _TITLE_FONT
    c.alignment = Alignment(vertical="center")

    ws["A2"] = f"대상 파일: {os.path.basename(result.source_path)}"
    ws["A3"] = f"점검 일자: {date.today().isoformat()}    /    점검 방식: {config.REVIEW_METHOD}"
    ws["A4"] = (
        f"점검대상 {applicable}개 중  적합 {counts.get('적합', 0)} · "
        f"보완필요 {counts.get('보완필요', 0)} · 누락 {counts.get('누락', 0)}   "
        f"(해당없음 {counts.get('해당없음', 0)}개 제외)"
    )
    ws["A5"] = f"총평: {result.overall_summary}"
    for r in (2, 3, 4, 5):
        ws.merge_cells(f"A{r}:{last_col}{r}")
        ws[f"A{r}"].font = _META_FONT
        ws[f"A{r}"].alignment = _WRAP_TOP
    return 7  # 표 헤더 행


def _write_table(ws, result: ReviewResult, header_row: int, cidx: dict) -> None:
    # 헤더
    for col, (title, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = f"A{header_row + 1}"

    # findings를 점검표 순서로 정렬(미등록 항목은 뒤로)
    def sort_key(f):
        meta = cidx.get(f.item_id)
        return meta["order"] if meta else 9999

    findings = sorted(result.findings, key=sort_key)
    row = header_row + 1
    if not findings:
        ws.cell(row=row, column=1, value="-")
        ws.cell(row=row, column=2, value="(점검 결과 없음 — 빈 문서 또는 판단 누락)")
        for col in range(1, _NCOL + 1):
            ws.cell(row=row, column=col).border = _BORDER
            ws.cell(row=row, column=col).alignment = _WRAP_TOP
        return

    for f in findings:
        meta = cidx.get(f.item_id, {})
        values = [
            meta.get("area_no", ""),
            meta.get("area", ""),
            f.item_name or meta.get("item_name", f.item_id),
            f.verdict,
            f.found_content,
            f.issue,
            f.recommendation,
            f.confidence,
            f.legal_basis or meta.get("law_basis", ""),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _BORDER
            cell.alignment = _CENTER if col in (1, 4, 8) else _WRAP_TOP
        ws.cell(row=row, column=4).fill = _VERDICT_FILL.get(f.verdict, PatternFill())
        ws.cell(row=row, column=8).fill = _CONF_FILL.get(f.confidence, PatternFill())
        row += 1


def _write_summary(ws, results: list[ReviewResult]) -> None:
    headers = [("파일", 30), ("적합", 7), ("보완필요", 9), ("누락", 7),
               ("해당없음", 9), ("총평", 70)]
    for col, (title, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    for i, result in enumerate(results, start=2):
        counts = Counter(f.verdict for f in result.findings)
        values = [
            os.path.basename(result.source_path),
            counts.get("적합", 0),
            counts.get("보완필요", 0),
            counts.get("누락", 0),
            counts.get("해당없음", 0),
            result.overall_summary,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = _BORDER
            cell.alignment = _CENTER if 2 <= col <= 5 else _WRAP_TOP


def write_report(results: list[ReviewResult], out_path: str) -> None:
    """점검 결과를 엑셀 파일로 저장한다(요약 시트 + 문서별 상세 시트)."""
    cidx = _checklist_index()
    wb = Workbook()
    summary = wb.active
    summary.title = "요약"
    _write_summary(summary, results)

    used_names = {"요약"}
    for result in results:
        counts = Counter(f.verdict for f in result.findings)
        applicable = sum(c for v, c in counts.items() if v != "해당없음")
        sheet_name = _safe_sheet_name(
            os.path.splitext(os.path.basename(result.source_path))[0], used_names
        )
        ws = wb.create_sheet(sheet_name)
        header_row = _write_block(ws, result, counts, applicable)
        _write_table(ws, result, header_row, cidx)

    out_dir = os.path.dirname(os.path.abspath(out_path))
    os.makedirs(out_dir, exist_ok=True)
    wb.save(out_path)
